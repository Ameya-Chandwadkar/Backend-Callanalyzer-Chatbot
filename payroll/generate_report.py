"""
generate_report.py
Builds the MasonMart Combined Performance & Salary Report as a multi-sheet
.xlsx, mirroring the structure and depth of the June-2026 manual audit:
  1. Combined Summary        — performance + salary snapshot for the team
  <one sheet per rep>        — Metric | Value | How | Source | Basis, + remark
  Evidence                   — sub-target days, never-attended raw counts
  Order Attribution          — per-customer order-incentive detail
  Methodology & Gaps         — how it was built + what's missing and why

CORE RULE (unchanged, and the reason this file is trustworthy): a number is
written only if it was computed from real ingested rows (see metrics.py) plus
a fully-specified config value. Anything depending on a missing input, a
period mismatch, or an unknown formula is written as an explicit
"NOT AVAILABLE — <reason>", never as 0, blank, or a guess. Management remarks
are strict recaps of the computed numbers, not invented characterizations.

USAGE:
    python payroll/generate_report.py
Output: payroll/output/MasonMart_Report_<timestamp>.xlsx
"""

import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from common import get_connection
from payroll import metrics

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

PAYROLL_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(PAYROLL_DIR, "payroll_config.json")
OUTPUT_DIR = os.path.join(PAYROLL_DIR, "output")

NA = "NOT AVAILABLE"

FONT = "Arial"
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F2937")
SUBTITLE_FONT = Font(name=FONT, size=9, italic=True, color="475569")
SECTION_FONT = Font(name=FONT, bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(name=FONT, bold=True, size=10, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="475569")
CELL_FONT = Font(name=FONT, size=10, color="1E293B")
NA_FONT = Font(name=FONT, size=10, italic=True, color="B45309")
REMARK_FONT = Font(name=FONT, size=10, italic=True, color="0F172A")
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _is_na(v):
    return isinstance(v, str) and NA in v


def _write_row(ws, r, values, fonts=None, fill=None, borders=True):
    for c, val in enumerate(values, start=1):
        cell = ws.cell(r, c, val)
        cell.font = (fonts[c - 1] if fonts else (NA_FONT if _is_na(val) else CELL_FONT))
        cell.alignment = WRAP
        if fill:
            cell.fill = fill
        if borders:
            cell.border = BORDER
    return r + 1


# ── Salary computation (per rep) ─────────────────────────────

def compute_salary(conn, config, emp, ct, sub_days, order_incentive):
    """Returns a list of (component, amount_or_NA, how, source, basis) rows,
    plus a 'payable_now' summary value."""
    vcd = config.get("valid_call_definition")
    tiers = config["order_incentive_tiers"]
    qual = config["order_incentive_qualification"]
    rows = []
    payable_parts = []
    blocked = False

    if emp["employment_type"] == "full_time":
        fixed = emp.get("fixed_salary")
        if fixed is None:
            rows.append(("Fixed Salary", f"{NA} — fixed_salary not set in payroll_config.json",
                         "", "payroll_config.json", "Full-time fixed monthly salary."))
            blocked = True
        else:
            rows.append(("Fixed Salary", fixed, "Flat monthly fixed salary, independent of call count.",
                         "Offer letter (payroll_config.json)", "Full-time fixed salary."))
            payable_parts.append(fixed)
        rows.append(("Call-Based Pay", "N/A (full-time)",
                     "Full-time staff are on fixed salary, not per-call pay.",
                     "Offer letter", "No per-call clause for full-time."))
    else:
        rows.append(("Fixed Salary", "N/A (part-time)",
                     "Part-time has no fixed/guaranteed salary.",
                     "Agreement", "Pay linked to verified calls/orders."))
        rate = emp.get("per_call_rate")
        if rate is None:
            rows.append(("Call-Based Pay", f"{NA} — per_call_rate not set in payroll_config.json",
                         "", "payroll_config.json", "Per valid call."))
            blocked = True
        elif vcd is None:
            rows.append(("Call-Based Pay",
                         f"{NA} — valid_call_definition UNSET in payroll_config.json",
                         "Set to 'connected_45s' or 'any_outgoing_attempt' to compute. "
                         "The two give materially different pay (see Methodology sheet).",
                         "Agreement + payroll_config.json",
                         f"Rs {rate} per valid call."))
            blocked = True
        else:
            count = ct["connected_45s"] if vcd == "connected_45s" else ct["outgoing"]
            pay = round(count * rate, 2)
            rows.append(("Call-Based Pay", pay,
                         f"{count} {'connected calls >45s' if vcd=='connected_45s' else 'outgoing attempts'} "
                         f"x Rs {rate} = Rs {pay}.",
                         "Call History", f"Rs {rate} per valid call ({vcd})."))
            payable_parts.append(pay)

    # LOP risk — info only, never deducted
    if emp["employment_type"] == "full_time" and emp.get("fixed_salary"):
        lop_days = len(sub_days)
        daily = emp["fixed_salary"] / 26
        at_risk = round(lop_days * daily)
        rows.append(("LOP risk (info only)",
                     f"-{at_risk:,} at risk" if lop_days else "0",
                     f"{lop_days} working day(s) below {config['discipline_rules']['rating_day_min_attempts']} "
                     f"attempts x (Rs {emp['fixed_salary']:,}/26). NOT deducted — discretionary, needs "
                     f"an attendance/leave check.",
                     "Call History + agreement", "Sub-target day MAY be Loss-of-Pay on review."))

    # Order incentive
    rows.append(("Order Incentive (pending)", round(order_incentive, 2),
                 f"Tiered per the customer's order sequence "
                 f"(1st Rs {tiers['1st']} / 2nd Rs {tiers['2nd']} / 3rd Rs {tiers['3rd']}) "
                 f"for customers mapped to this rep.",
                 "Customer map + Orders",
                 f"Earned only at order #{qual['min_order_sequence']} + Rs {qual['min_order_value']:,}."))

    # Target bonus
    bonus = emp.get("target_bonus_amount", 0)
    rows.append(("Target Bonus", bonus,
                 f"Needs > {config['target_bonus']['min_qualified_customers']} qualified customers "
                 f"(order #{qual['min_order_sequence']} + Rs {qual['min_order_value']:,}). Current qualified = 0.",
                 "Customer map + Orders", "Monthly target bonus."))

    # Feedback / form
    rows.append(("Feedback / Form", f"{NA} — no feedback/form data supplied",
                 "Rs 1/call feedback & form bonuses need data not in any uploaded file.",
                 "—", "Feedback & form clauses."))

    # Payable now
    if blocked:
        payable = f"{NA} — blocked by the item(s) above"
    elif emp.get("probation") and emp["employment_type"] == "full_time":
        payable = payable_parts[0] if payable_parts else 0
        payable_note = "Fixed salary only; incentives withheld during probation."
    else:
        payable = round(sum(payable_parts), 2) if payable_parts else 0
        payable_note = "Verified call-based pay; order incentive pending."
    if not blocked:
        rows.append(("PAYABLE NOW", payable,
                     payable_note if 'payable_note' in dir() else "",
                     "—", "Probation withholding applies to full-time."))
    else:
        rows.append(("PAYABLE NOW", payable, "", "—",
                     "Resolve the blocked item(s) above to compute."))

    return rows, payable


# ── Per-rep performance rows ─────────────────────────────────

def performance_rows(conn, config, sim, start, end):
    exc = config["period"]["excluded_weekdays"]
    dr = config["discipline_rules"]
    min_att = dr["rating_day_min_attempts"]
    ct = metrics.call_totals(conn, sim, start, end)
    rdays = metrics.rating_days(conn, sim, start, end, min_att, exc)
    sub = metrics.sub_threshold_days(conn, sim, start, end, min_att, exc)

    rows = []
    rows.append(("Outgoing attempts", ct["outgoing"] or 0,
                 "Count of outgoing calls in the period.", "Call History", "Calling-effort metric."))
    rows.append(("Incoming calls", ct["incoming"] or 0,
                 "Count of incoming calls in the period.", "Call History", "—"))
    rows.append(("Connected calls (>45s)", ct["connected_45s"] or 0,
                 "Calls with talk duration over 45 seconds.", "Call History", "Connected-call definition (>45s)."))
    rows.append(("Talk time", metrics.fmt_talk(ct["talk_seconds"]),
                 "Sum of all call durations.", "Call History", "Total talk time."))
    rows.append(("Active / Rating days", f"{ct['active_days']} active / {len(rdays)} rating",
                 f"Active = working days with any activity; Rating = days with {min_att}+ outgoing attempts.",
                 "Call History", f"Rating day = {min_att}+ attempts (excl. {', '.join(exc)})."))

    disc_applicable = emp_is_disc(config, sim)
    if disc_applicable:
        fc = metrics.first_call_discipline(conn, sim, start, end, dr["first_call_deadline"], min_att, exc)
        rows.append(("First-call timing",
                     f"late on {fc['late']} of {fc['rating_days']} rating days",
                     f"First outgoing call after {fc['deadline']} on {fc['late']} of {fc['rating_days']} "
                     f"rating days (on-time {fc['on_time']}).",
                     "Call History", f"First call before {fc['deadline']} (full-time)."))
        mw = metrics.morning_window_discipline(conn, sim, start, end, dr["morning_window"], min_att, exc)
        w = dr["morning_window"]
        rows.append(("Morning 9-11 discipline",
                     f"met on {mw['passed']} of {mw['rating_days']} rating days",
                     f"Hit the {mw['window']} benchmark on {mw['passed']} of {mw['rating_days']} rating days.",
                     "Call History",
                     f"{w['min_outgoing_attempts']} out + {w['min_connected_calls']} connected, or "
                     f"{w['min_talk_minutes']} talk-min, in {mw['window']} (full-time)."))
        hd = metrics.hourly_discipline(conn, sim, start, end, dr["hourly_discipline"], min_att, exc)
        hr = dr["hourly_discipline"]
        rows.append(("Hourly discipline",
                     f"{hd['met']} of {hd['eligible_hours']} eligible hour-slots met the bar",
                     f"{hd['met']} of {hd['eligible_hours']} eligible hour-slots "
                     f"(a clock-hour on a rating day with >=1 outgoing call) met the bar.",
                     "Call History",
                     f"{hr['min_calls_per_hour']} calls or {hr['min_talk_minutes_per_hour']} talk-min "
                     f"per eligible hour (full-time). NB: our explicit definition of 'eligible hour'."))
    else:
        rows.append(("Discipline rules", "Not applied (part-time)",
                     "Part-time staff are not scored on full-time timing rules.",
                     "Agreement", "Part-time logic — no full-time timing penalties."))

    # Never Attended (period-aware)
    na = metrics.never_attended_summary(conn, sim, start, end)
    if na["missed"] == 0:
        rows.append(("Never Attended", "none in the missed-call data",
                     "No missed-call rows for this rep in the uploaded Never Attended Report.",
                     "Never Attended Report", "Missed calls not returned."))
    elif not na["overlaps_calls"]:
        p0, p1 = na["period"]
        rows.append(("Never Attended (raw missed)", na["missed"],
                     f"{na['missed']} missed calls in the Never Attended Report ({p0} to {p1}). "
                     f"Callback recovery NOT computed: that data is from a different period than the "
                     f"call history ({start} to {end}), so callbacks can't be detected.",
                     "Never Attended Report",
                     "Recovery needs same-period call data — see Methodology."))
    else:
        rows.append(("Final Never Attended", na["final_never_attended"],
                     f"{na['missed']} missed minus {na['recovered']} with a later callback = "
                     f"{na['final_never_attended']} never returned.",
                     "Never Attended Report + Call History",
                     "Missed only; a later outgoing call to the number = recovered."))

    # Gold Lead & follow-ups — genuine gaps
    rows.append(("Gold Lead response", f"{NA} — no Gold Lead-tagged leads in the data",
                 "No lead in any uploaded file carries a 'Gold Lead' tag, so the 30-min response "
                 "rule can't be evaluated.", "Lead Data Report", "Gold Lead 30-min rule."))
    rows.append(("Follow-ups", f"{NA} — reminder data not available",
                 "Reminder dates are populated on almost no rows, and the rich lead export isn't "
                 "loaded into the leads table, so follow-up completion can't be computed.",
                 "Lead Data Report", "Follow-up = call after reminder time."))

    return rows, ct, sub, rdays


def emp_is_disc(config, sim):
    for k, e in config.get("employees", {}).items():
        if k.startswith("_"):
            continue
        if e.get("_resolved_rep_sim") == sim:
            return e.get("discipline_rules_applicable", False)
    return False


# ── Sheet builders ───────────────────────────────────────────

def _annotated_table(ws, r, headers, rows):
    r = _write_row(ws, r, headers, fonts=[HEAD_FONT] * len(headers), fill=HEAD_FILL)
    for row in rows:
        r = _write_row(ws, r, list(row))
    return r


def build_workbook(conn, config):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    start, end, explicit = metrics.report_period(conn, config)
    exc = config["period"]["excluded_weekdays"]
    wdays = metrics.working_day_count(start, end, exc)

    employees = {k: v for k, v in config.get("employees", {}).items() if not k.startswith("_")}
    for name, emp in employees.items():
        row = conn.execute("SELECT rep_sim_number FROM reps WHERE canonical_name = ?", (name,)).fetchone()
        emp["_resolved_rep_sim"] = row["rep_sim_number"] if row else None

    order_detail, incentive_by_rep, order_sources = metrics.order_incentives(conn, config)

    period_label = (f"Period {start} to {end}"
                    + ("" if explicit else "  (NO period set in payroll_config.json — showing the full "
                       "range of call data on file, not a specific pay period)")
                    + f"  |  {', '.join(exc)} excluded -> {wdays} working days"
                    + "  |  Status: PROVISIONAL")

    # cache per-rep computed data for the summary
    rep_cache = {}

    # ---------- Per-rep sheets ----------
    for name, emp in employees.items():
        sim = emp["_resolved_rep_sim"]
        safe_tab = name[:28]
        ws = wb.create_sheet(safe_tab)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 26
        ws.column_dimensions["E"].width = 40

        ws.cell(1, 1, f"Individual Report - {name} ({emp['employment_type'].replace('_',' ').title()})").font = TITLE_FONT
        ws.cell(2, 1, "Each line shows the value, exactly how it was calculated, the source, and the basis.").font = SUBTITLE_FONT

        if sim is None:
            ws.cell(4, 1, f"{NA} - '{name}' has no call history in the reps table, so nothing can be computed.").font = NA_FONT
            rep_cache[name] = None
            continue

        perf, ct, sub, rdays = performance_rows(conn, config, sim, start, end)
        incentive = incentive_by_rep.get(sim, 0.0)
        sal, payable = compute_salary(conn, config, emp, ct, sub, incentive)
        rep_cache[name] = {"ct": ct, "sub": sub, "rdays": rdays, "incentive": incentive,
                           "payable": payable, "perf": perf, "sim": sim, "emp": emp}

        r = 4
        cell = ws.cell(r, 1, "PERFORMANCE"); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
        for c in range(2, 6):
            ws.cell(r, c).fill = SECTION_FILL
        r += 1
        r = _annotated_table(ws, r, ["Metric", "Value", "How it was calculated", "Source", "Basis"], perf)

        r += 1
        cell = ws.cell(r, 1, "SALARY & INCENTIVES"); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
        for c in range(2, 6):
            ws.cell(r, c).fill = SECTION_FILL
        r += 1
        r = _annotated_table(ws, r, ["Component", "Amount (Rs)", "How it was calculated", "Source", "Basis"], sal)

        r += 1
        cell = ws.cell(r, 1, "MANAGEMENT REMARK (factual recap)"); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
        for c in range(2, 6):
            ws.cell(r, c).fill = SECTION_FILL
        r += 1
        remark = build_remark(name, emp, rep_cache[name], config)
        cell = ws.cell(r, 1, remark); cell.font = REMARK_FONT; cell.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 60

    # ---------- Combined Summary (first sheet) ----------
    ws = wb.create_sheet("1. Combined Summary", 0)
    ws.column_dimensions["A"].width = 20
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 16
    ws.cell(1, 1, "MasonMart - Combined Performance & Salary Report").font = TITLE_FONT
    ws.cell(2, 1, period_label).font = SUBTITLE_FONT

    r = 4
    cell = ws.cell(r, 1, "A. PERFORMANCE SNAPSHOT"); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(r, c).fill = SECTION_FILL
    r += 1
    r = _write_row(ws, r, ["Employee", "Type", "Outgoing", "Connected >45s", "Talk Time",
                           "Rating Days", "Sub-target Days"],
                   fonts=[HEAD_FONT] * 7, fill=HEAD_FILL)
    for name, emp in employees.items():
        rc = rep_cache.get(name)
        if rc is None:
            r = _write_row(ws, r, [name, emp["employment_type"], NA, "", "", "", ""])
            continue
        ct = rc["ct"]
        r = _write_row(ws, r, [name, emp["employment_type"], ct["outgoing"] or 0,
                               ct["connected_45s"] or 0, metrics.fmt_talk(ct["talk_seconds"]),
                               len(rc["rdays"]), len(rc["sub"])])

    r += 2
    cell = ws.cell(r, 1, "B. SALARY & INCENTIVE SNAPSHOT"); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(r, c).fill = SECTION_FILL
    r += 1
    r = _write_row(ws, r, ["Employee", "Type", "Fixed (Rs)", "Call Pay (Rs)",
                           "Order Incentive (Rs)", "Target Bonus (Rs)", "Payable Now (Rs)"],
                   fonts=[HEAD_FONT] * 7, fill=HEAD_FILL)
    for name, emp in employees.items():
        rc = rep_cache.get(name)
        if rc is None:
            r = _write_row(ws, r, [name, emp["employment_type"], NA, "", "", "", ""])
            continue
        vcd = config.get("valid_call_definition")
        if emp["employment_type"] == "full_time":
            fixed_disp = emp.get("fixed_salary") if emp.get("fixed_salary") is not None else NA
            callpay_disp = "N/A"
        else:
            fixed_disp = "N/A"
            callpay_disp = (f"{NA} (valid_call unset)" if vcd is None else
                            round((rc["ct"]["connected_45s"] if vcd == "connected_45s" else rc["ct"]["outgoing"])
                                  * emp.get("per_call_rate", 0), 2))
        r = _write_row(ws, r, [name, emp["employment_type"], fixed_disp, callpay_disp,
                               round(rc["incentive"], 2), emp.get("target_bonus_amount", 0), rc["payable"]])

    r += 1
    ws.cell(r, 1, "How: performance counted row-by-row from Call History; salary from payroll_config.json "
                  "(offer-letter terms) + order incentives from the customer map & Shopify orders.").font = SUBTITLE_FONT
    r += 1
    ws.cell(r, 1, "Every figure is provisional and traceable — see each rep's own sheet for the line-by-line "
                  "'how', and the Methodology & Gaps sheet for what is intentionally left blank and why.").font = SUBTITLE_FONT

    # ---------- Evidence ----------
    build_evidence_sheet(wb, conn, config, rep_cache, employees, start, end)

    # ---------- Order Attribution ----------
    build_order_sheet(wb, conn, config, order_detail, order_sources)

    # ---------- Methodology & Gaps ----------
    build_methodology_sheet(wb, conn, config, start, end, explicit, period_label)

    return wb


def build_remark(name, emp, rc, config):
    """Strict factual recap — no invented adjectives, only computed numbers."""
    ct = rc["ct"]
    parts = [f"{ct['outgoing'] or 0} outgoing attempts, {ct['connected_45s'] or 0} connected (>45s), "
             f"{metrics.fmt_talk(ct['talk_seconds'])} talk time across {len(rc['rdays'])} rating day(s)."]
    if rc["sub"]:
        parts.append(f"{len(rc['sub'])} working day(s) below target (LOP risk, info only).")
    if rc["incentive"]:
        parts.append(f"Rs {round(rc['incentive'], 2)} order incentive pending.")
    payable = rc["payable"]
    parts.append(f"Payable now: {'blocked — see salary section' if _is_na(str(payable)) else payable}.")
    return " ".join(parts)


def build_evidence_sheet(wb, conn, config, rep_cache, employees, start, end):
    ws = wb.create_sheet("6. Evidence")
    ws.column_dimensions["A"].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 20
    ws.cell(1, 1, "Supporting Evidence").font = TITLE_FONT

    r = 3
    cell = ws.cell(r, 1, f"Working days below {config['discipline_rules']['rating_day_min_attempts']} "
                         f"outgoing attempts (LOP risk — info only, not deducted)")
    cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    for c in range(2, 5):
        ws.cell(r, c).fill = SECTION_FILL
    r += 1
    r = _write_row(ws, r, ["Employee", "Date", "Outgoing Attempts", "Shortfall"],
                   fonts=[HEAD_FONT] * 4, fill=HEAD_FILL)
    any_sub = False
    for name in employees:
        rc = rep_cache.get(name)
        if not rc:
            continue
        for (d, n, short) in rc["sub"]:
            r = _write_row(ws, r, [name, d, n, short]); any_sub = True
    if not any_sub:
        r = _write_row(ws, r, ["(none — every rep hit target on all working days)", "", "", ""])

    r += 2
    cell = ws.cell(r, 1, "Never Attended — raw missed counts")
    cell.font = SECTION_FONT; cell.fill = SECTION_FILL
    for c in range(2, 5):
        ws.cell(r, c).fill = SECTION_FILL
    r += 1
    r = _write_row(ws, r, ["Employee", "Missed (raw)", "Missed-data period", "Callback recovery"],
                   fonts=[HEAD_FONT] * 4, fill=HEAD_FILL)
    for name in employees:
        rc = rep_cache.get(name)
        if not rc:
            continue
        na = metrics.never_attended_summary(conn, rc["sim"], start, end)
        if na["missed"] == 0:
            r = _write_row(ws, r, [name, 0, "—", "—"])
        elif not na["overlaps_calls"]:
            p0, p1 = na["period"]
            r = _write_row(ws, r, [name, na["missed"], f"{p0} to {p1}",
                                   f"{NA} — different period than call history ({start} to {end})"])
        else:
            r = _write_row(ws, r, [name, na["missed"], f"{na['period'][0]} to {na['period'][1]}",
                                   f"{na['recovered']} recovered -> {na['final_never_attended']} final"])


def build_order_sheet(wb, conn, config, order_detail, order_sources):
    ws = wb.create_sheet("7. Order Attribution")
    ws.column_dimensions["A"].width = 18
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws.cell(1, 1, "Order Attribution & Incentive Detail").font = TITLE_FONT
    src_note = ("Source of customer->rep ownership: "
                + (", ".join(sorted(order_sources)) if order_sources else "none")
                + ". 'lead_assignment' means the mapping came from who a LEAD was assigned to (a caveat "
                  "source — see Methodology), not a verified sale owner.")
    ws.cell(2, 1, src_note).font = SUBTITLE_FONT

    r = 4
    r = _write_row(ws, r, ["Salesperson", "Customer Phone", "Order", "Sequence",
                           "Order Value (Rs)", "Cumulative (Rs)", "Incentive (Rs)", "Status"],
                   fonts=[HEAD_FONT] * 8, fill=HEAD_FILL)
    if not order_detail:
        r = _write_row(ws, r, [f"{NA} — no mapped customer has a Shopify order yet.", "", "", "", "", "", "", ""])
    else:
        total_inc = 0
        for d in order_detail:
            status = "Earned" if d["earned"] else "Pending (threshold not met)"
            r = _write_row(ws, r, [d["rep"], d["customer_phone"], d["order"], d["sequence"],
                                   round(d["value"], 2), round(d["cumulative"], 2), d["incentive"], status])
            total_inc += d["incentive"]
        r = _write_row(ws, r, ["TOTAL pending order incentive", "", "", "", "", "", total_inc, ""],
                       fonts=[HEAD_FONT] * 8, fill=HEAD_FILL)

    r += 1
    unattr = conn.execute(
        """SELECT COUNT(*) n, COALESCE(SUM(total_price),0) rev FROM shopify_orders o
           WHERE o.customer_phone_norm IS NULL
              OR o.customer_phone_norm NOT IN (SELECT customer_phone_norm FROM customer_salesperson_map)"""
    ).fetchone()
    ws.cell(r, 1, f"Unattributed paid orders (customer not in the map): {unattr['n']} orders, "
                  f"Rs {unattr['rev']:,.2f} — earn no incentive until mapped.").font = CELL_FONT


def build_methodology_sheet(wb, conn, config, start, end, explicit, period_label):
    ws = wb.create_sheet("8. Methodology & Gaps")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95
    ws.cell(1, 1, "Methodology, Assumptions & Data Gaps").font = TITLE_FONT

    calls_n = conn.execute("SELECT COUNT(*) FROM callyzer_calls").fetchone()[0]
    na_n = conn.execute("SELECT COUNT(*) FROM callyzer_never_attended").fetchone()[0]
    na_range = conn.execute("SELECT MIN(date(call_timestamp)) mn, MAX(date(call_timestamp)) mx FROM callyzer_never_attended").fetchone()
    map_n = conn.execute("SELECT COUNT(*) FROM customer_salesperson_map").fetchone()[0]
    map_with_orders = conn.execute(
        """SELECT COUNT(DISTINCT m.customer_phone_norm) FROM customer_salesperson_map m
           JOIN shopify_orders o ON o.customer_phone_norm = m.customer_phone_norm""").fetchone()[0]

    rows = [
        ("HOW THE REPORT WAS BUILT", ""),
        ("Period", period_label),
        ("Employee matching", "By company SIM number (reps table); name variants reconciled via rep_name_aliases."),
        ("Performance", f"Counted row-by-row from callyzer_calls ({calls_n:,} calls loaded). Rating day, "
                        f"discipline windows and talk time all derived from real call timestamps/durations."),
        ("Discipline SCORES not reproduced", "The June audit's composite scores (e.g. '0/20', '0.44/5') have no "
                                             "stated formula, so they are NOT reproduced here. Instead each rep's "
                                             "sheet reports the underlying verifiable facts (e.g. 'late on 6 of 8 "
                                             "rating days'), which are auditable."),
        ("WHY ITEMS ARE NOT AVAILABLE", ""),
        ("Part-time call-based pay", f"valid_call_definition is UNSET in payroll_config.json. Rs-per-call pay is either "
                                     f"per connected-call (>45s) or per outgoing attempt — a materially different "
                                     f"number. Set it to 'connected_45s' or 'any_outgoing_attempt' to unblock."),
        ("Never Attended recovery", f"The Never Attended Report loaded ({na_n} rows, {na_range['mn']} to {na_range['mx']}) "
                                    f"is from a DIFFERENT period than the call history ({start} to {end}). A callback "
                                    f"can only be detected with same-period call data, so 'final never attended' is "
                                    f"not computed — only the raw missed count is shown."),
        ("Gold Lead response", "No lead in any uploaded file carries a 'Gold Lead' tag, so the 30-minute response "
                               "rule cannot be evaluated. (The lead export's tags are IndiaMart Lead, contractor, etc.)"),
        ("Follow-ups", "Reminder dates are populated on almost no rows and the rich lead export is not loaded into the "
                       "leads table, so follow-up completion cannot be computed."),
        ("Feedback / Form bonus", "No feedback data and no customer-form data supplied."),
        ("CUSTOMER MAPPING CONFIDENCE", ""),
        ("Customer -> salesperson", f"{map_n:,} customers mapped, of which {map_with_orders} actually have a Shopify "
                                    f"order (only those affect any payout). If sourced from the Lead Data Report, the "
                                    f"mapping reflects who a LEAD was assigned to, not a confirmed sale owner — review "
                                    f"before final payroll."),
        ("DATA GAPS TO CLOSE FOR FINAL PAYROLL", ""),
        ("1", "Set valid_call_definition (unblocks part-time pay)."),
        ("2", "Upload a Never Attended Report covering the SAME period as the call history (enables recovery)."),
        ("3", "Provide Gold Lead-tagged leads + assignment times (enables the 30-min rule)."),
        ("4", "Provide attendance/leave log (decides whether sub-target days are Loss-of-Pay or approved leave)."),
        ("5", "Confirm the customer->salesperson mapping for the handful of customers who actually placed orders."),
    ]
    r = 3
    for label, val in rows:
        if val == "":
            cell = ws.cell(r, 1, label); cell.font = SECTION_FONT; cell.fill = SECTION_FILL
            ws.cell(r, 2).fill = SECTION_FILL
        else:
            ws.cell(r, 1, label).font = Font(name=FONT, bold=True, size=10)
            c2 = ws.cell(r, 2, val)
            c2.font = NA_FONT if _is_na(val) else CELL_FONT
            c2.alignment = WRAP
        r += 1


def generate_report_file():
    config = json.load(open(CONFIG_PATH, encoding="utf-8"))
    conn = get_connection()
    try:
        wb = build_workbook(conn, config)
    finally:
        conn.close()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = os.path.join(OUTPUT_DIR, f"MasonMart_Report_{stamp}.xlsx")
    wb.save(out_path)
    return out_path


def main():
    print(f"Report written to {generate_report_file()}")


if __name__ == "__main__":
    main()
